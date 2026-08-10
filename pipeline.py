#!/usr/bin/env python3
"""
Núcleo de consolidação DETALHADA (Balanço + DRE + DFC) a partir de balancetes
de verificação — de uma única empresa ou de várias empresas de um mesmo
grupo (soma simples, sem eliminação de saldos intercompanhias). Módulo
autocontido, usado pelo app Streamlit (app.py) e reaproveitável em scripts de
linha de comando.

IMPORTANTE — suposição de plano de contas: os códigos de classificação usados
abaixo (ex.: '1.1.01' = Caixa, '2.1.03' = Fornecedores, '3.1' = Custos,
'4.1.01' = Receita Bruta) foram validados contra os balancetes e o relatório
consolidado oficial do Grupo CLA/Claudio. Eles tendem a generalizar bem para
outras empresas que usem uma estrutura de plano de contas semelhante (comum
quando o mesmo escritório de contabilidade padroniza a codificação entre seus
clientes), mas NÃO são garantidos para qualquer empresa. Ao usar com uma
empresa/grupo novo pela primeira vez, sempre confira a tabela de "Checagem de
consistência" (deve ficar em ~0,00) e, se possível, compare pelo menos um
período contra o Balanço/DRE oficial dessa empresa antes de confiar no
resultado.
"""
import glob
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 0. Núcleo de leitura de balancete + constantes de formatação
# --------------------------------------------------------------------------

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
GRANDTOTAL_FILL = PatternFill("solid", fgColor="C6E0B4")
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
NUMFMT = '#,##0.00;(#,##0.00);"-"'


def normalize_desc(s):
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s.strip().upper()


def _parse_saldo(v):
    """Converte um valor de saldo para float, aceitando tanto número puro
    (formato de balancete onde o sinal já vem embutido, ex.: -595249.17 para
    saldo credor) quanto texto no formato brasileiro com sufixo D/C no final
    (ex.: "4.122.484,29D" = devedor/positivo, "511.575,35C" = credor/negativo
    — visto em balancetes de outros sistemas/escritórios de contabilidade)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        sign = 1.0
        if s[-1] in ('D', 'd'):
            s = s[:-1].strip()
        elif s[-1] in ('C', 'c'):
            s = s[:-1].strip()
            sign = -1.0
        s = s.replace('.', '').replace(',', '.')
        try:
            return sign * float(s)
        except ValueError:
            return None
    return None


def read_balancete_rows(ws):
    """Retorna lista de (codigo_classificacao, descricao, saldo_atual, saldo_anterior)."""
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row and any(isinstance(c, str) and 'Classificação' in c for c in row if c):
            header_row = i
            break
    if header_row is None:
        return []
    header = [c for c in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)][0]
    col_class = next((i for i, c in enumerate(header) if isinstance(c, str) and 'Classificação' in c), None)
    col_desc = next((i for i, c in enumerate(header) if isinstance(c, str) and 'Descrição' in c), None)
    col_sa = next((i for i, c in enumerate(header) if isinstance(c, str) and 'Saldo Anterior' in c), None)
    col_sat = next((i for i, c in enumerate(header) if isinstance(c, str) and 'Saldo Atual' in c), None)
    if col_class is None or col_desc is None or col_sat is None:
        return []
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if col_class >= len(row) or col_desc >= len(row) or col_sat >= len(row):
            continue
        code, desc = row[col_class], row[col_desc]
        if not isinstance(code, str) or not isinstance(desc, str):
            continue
        sat = _parse_saldo(row[col_sat])
        if sat is None:
            continue
        sa = None
        if col_sa is not None and col_sa < len(row):
            sa = _parse_saldo(row[col_sa])
        out.append((code.strip(), desc.strip(), sat, sa))
    return out


def get_code_value(balancete_rows, code, field=2):
    for row in balancete_rows:
        if row[0] == code:
            val = row[field]
            return val if val is not None else 0.0
    return 0.0


def find_leaf_by_desc(balancete_rows, code_prefix, keywords, exclude_prefixes=()):
    kws = [normalize_desc(k) for k in keywords]
    cands = []
    for code, desc, sat, sa in balancete_rows:
        if not code.startswith(code_prefix + '.'):
            continue
        if any(code.startswith(ex) for ex in exclude_prefixes):
            continue
        nd = normalize_desc(desc)
        if all(k in nd for k in kws):
            cands.append((code, sat))
    cands.sort(key=lambda t: len(t[0]))
    kept_codes, total = [], 0.0
    for code, sat in cands:
        if any(code.startswith(kc + '.') for kc in kept_codes):
            continue
        kept_codes.append(code)
        total += sat
    return total


def find_leaf_by_desc_field(bal_rows, code_prefix, keywords, field=2, exclude_prefixes=()):
    kws = [normalize_desc(k) for k in keywords]
    cands = []
    for code, desc, sat, sa in bal_rows:
        if not code.startswith(code_prefix + '.'):
            continue
        if any(code.startswith(ex) for ex in exclude_prefixes):
            continue
        nd = normalize_desc(desc)
        if all(k in nd for k in kws):
            val = sat if field == 2 else (sa or 0.0)
            cands.append((code, val))
    cands.sort(key=lambda t: len(t[0]))
    kept_codes, total = [], 0.0
    for code, val in cands:
        if any(code.startswith(kc + '.') for kc in kept_codes):
            continue
        kept_codes.append(code)
        total += val
    return total


# --------------------------------------------------------------------------
# 1. Localizar e ler os balancetes de todos os períodos, por CNPJ
# --------------------------------------------------------------------------

def read_period(ws):
    """Aceita tanto o formato com rótulo e valor em células separadas
    ("Período:" | "01/01/2026 - 31/03/2026") quanto o formato com os dois
    juntos na mesma célula ("Período: 01/01/2026 - 30/06/2026" — visto em
    balancetes de outros sistemas/escritórios)."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for i, c in enumerate(row):
            if isinstance(c, str) and 'Período' in c:
                if i + 1 < len(row) and row[i + 1]:
                    return str(row[i + 1]).strip()
                m = re.search(r'Per[íi]odo\s*:?\s*(.+)', c, re.IGNORECASE)
                if m and m.group(1).strip():
                    return m.group(1).strip()
    return None


CNPJ_RE = re.compile(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}')

_LABEL_KEYWORDS = ['C.N.P.J', 'PERIODO', 'EMPRESA', 'INSC', 'CONSOLIDADO', 'BALANCETE',
                   'CLASSIFICACAO', 'CODIGO', 'DESCRICAO']


def _looks_like_label_row(text):
    nd = normalize_desc(text)
    return any(kw in nd for kw in _LABEL_KEYWORDS)


def _load_one(path):
    tmpdir = tempfile.mkdtemp(prefix="det_balancetes_")
    real_path = path
    if path.lower().endswith(".xls"):
        subprocess.run(["soffice", "--headless", "--convert-to", "xlsx",
                         "--outdir", tmpdir, path], capture_output=True, check=True)
        real_path = os.path.join(tmpdir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
    wb = load_workbook(real_path, data_only=True)
    ws = wb.worksheets[0]

    # CNPJ/Empresa/Período podem vir com rótulo e valor em células separadas
    # ("Empresa:" | "BRC TRANSPORTES LTDA" — formato de um sistema) ou com os
    # dois juntos numa única célula ("C.N.P.J.: 41.473.011/0001-76" — formato
    # de outro). Para o CNPJ, o mais robusto é procurar o padrão do número em
    # qualquer célula, independente de rótulo.
    cnpj, nome, cnpj_row_idx = None, None, None
    header_rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
    for ridx, row in enumerate(header_rows, start=1):
        for i, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            # \b evita confundir o rótulo "Empresa:" com a palavra "Empresas"
            # dentro de outro texto (ex.: "CONSOLIDADO (Empresas: 259,1332)").
            if re.search(r'\bEmpresa\b', cell, re.IGNORECASE):
                if i + 1 < len(row) and row[i + 1]:
                    nome = str(row[i + 1]).strip()
                else:
                    m = re.search(r'\bEmpresa\b\s*:?\s*(.+)', cell, re.IGNORECASE)
                    if m and m.group(1).strip():
                        nome = m.group(1).strip()
            m_cnpj = CNPJ_RE.search(cell)
            if m_cnpj:
                cnpj = m_cnpj.group(0)
                cnpj_row_idx = ridx
            elif 'C.N.P.J' in cell and i + 1 < len(row) and row[i + 1]:
                cnpj = str(row[i + 1]).strip()
                cnpj_row_idx = ridx

    # Se não achou rótulo "Empresa:", tenta a primeira célula de texto "solta"
    # (sem nenhum rótulo conhecido) antes da linha do CNPJ — é onde escritórios
    # que não rotulam o nome da empresa costumam colocá-lo (ex.: só o nome em
    # caixa alta na primeira linha da planilha).
    if not nome and cnpj_row_idx:
        for row in header_rows[:cnpj_row_idx - 1]:
            for cell in row:
                if isinstance(cell, str) and cell.strip() and not _looks_like_label_row(cell):
                    nome = cell.strip()
                    break
            if nome:
                break

    periodo = read_period(ws)
    bal_rows = read_balancete_rows(ws)
    return cnpj, nome, periodo, bal_rows


def load_all_balancetes(input_dir=None, file_paths=None):
    """Aceita uma pasta (input_dir) OU uma lista explícita de caminhos de
    arquivo (file_paths) — o app Streamlit usa file_paths (arquivos enviados
    pelo usuário e salvos num diretório temporário)."""
    if file_paths is None:
        file_paths = sorted(glob.glob(os.path.join(input_dir, "*.xls")) +
                             glob.glob(os.path.join(input_dir, "*.xlsx")))
    companies = {}
    warnings = []
    for x in file_paths:
        try:
            cnpj, nome, periodo, bal_rows = _load_one(x)
        except Exception as e:
            warnings.append(f"Falha ao ler {os.path.basename(x)}: {e}")
            continue
        if not cnpj:
            warnings.append(f"Não achei CNPJ em {os.path.basename(x)}, ignorado.")
            continue
        if not bal_rows:
            warnings.append(f"Não consegui ler linhas de balancete em {os.path.basename(x)} (cabeçalho 'Classificação' não encontrado?), ignorado.")
            continue
        companies.setdefault(cnpj, {'nome': nome, 'periodos': []})
        companies[cnpj]['periodos'].append((periodo, bal_rows, os.path.basename(x)))
    return companies, warnings


# --------------------------------------------------------------------------
# 2. Classificador por palavra-chave das "Despesas Gerais" (3.2.02.04)
# --------------------------------------------------------------------------

RULES = [
    ('depreciacao', ['DEPRECIA']), ('depreciacao', ['AMORTIZ']),
    ('transportes', ['VEICULO']), ('transportes', ['COMBUSTIVEL']),
    ('transportes', ['FRETE']), ('transportes', ['PEDAGIO']),
    ('transportes', ['TRANSITO']), ('transportes', ['CARRETO']),
    ('manutencao', ['MANUTEN']),
    ('informatica', ['INFORMATICA']), ('informatica', ['SISTEMA']),
    ('informatica', ['SOFTWARE']), ('informatica', ['TELECOMUNICA']),
    ('informatica', ['INTERNET']),
    ('consultoria', ['HONORARIO']), ('consultoria', ['CONTABIL']),
    ('consultoria', ['JURIDIC']), ('consultoria', ['ASSESSORIA']),
    ('consultoria', ['CONSULTORIA']), ('consultoria', ['LEGAIS']),
    ('consultoria', ['ASSISTENCIA']),
    ('prestacao_servicos', ['SERVICOS PRESTADOS']),
    ('marketing', ['PROPAGANDA']), ('marketing', ['PUBLICIDADE']),
    ('marketing', ['MARKETING']), ('marketing', ['PATROCINIO']),
    ('alugueis', ['ALUGUEL']),
    ('tributarias', ['ALVARA']),
    ('provisoes', ['PROVIS']),
    ('pessoal', ['CAPACITACAO']), ('pessoal', ['TREINAMENTO']),
]


def classify_desc(desc):
    nd = normalize_desc(desc)
    for cat, kws in RULES:
        if all(k in nd for k in kws):
            return cat
    return 'administrativas'


# --------------------------------------------------------------------------
# 2b. Resolução de grupos por PALAVRA-CHAVE da descrição, não por código fixo
# --------------------------------------------------------------------------
#
# Os balancetes de escritórios/ERPs diferentes numeram os grupos de forma
# diferente (ex.: Receita Bruta pode ser "4.1.01" num plano e "4.1.1" noutro;
# um cliente pode até inverter Receita<->Despesa entre os códigos 3 e 4). Em
# vez de confiar no número da conta, procuramos a linha cuja DESCRIÇÃO bate
# com o rótulo esperado (ex.: "ATIVO CIRCULANTE", "PASSIVO NÃO CIRCULANTE") —
# isso é muito mais estável entre empresas, porque a terminologia contábil em
# português é bem padronizada mesmo quando a numeração não é.


def _depth(code):
    return code.count('.') + 1


def find_group_row(bal_rows, keywords_all=None, keywords_any=None, depth=None,
                    exclude=(), code_prefix=None, field=2):
    """Acha a linha cuja descrição bate com as palavras-chave pedidas, dentro
    do escopo indicado (profundidade e/ou código pai), e devolve (codigo,
    valor) — o valor vem da MESMA linha que bateu, nunca de uma busca
    separada por código (alguns balancetes repetem o mesmo código de
    classificação em contas analíticas diferentes; buscar de novo por código
    poderia pegar o valor de uma conta errada). Quando mais de uma linha bate,
    usa a de MENOR profundidade — balancetes já trazem o subtotal calculado
    pelo ERP nessa linha, não precisamos somar as contas-filhas de novo."""
    keywords_all = [normalize_desc(k) for k in (keywords_all or [])]
    keywords_any = [normalize_desc(k) for k in (keywords_any or [])]
    exclude = [normalize_desc(k) for k in exclude]
    cands = []
    for code, desc, sat, sa in bal_rows:
        # quando um code_prefix é passado, estamos procurando um item DENTRO
        # desse grupo — nunca o próprio grupo (senão um rótulo de grupo tipo
        # "CONTAS DE RESULTADOS - CUSTOS E DESPESAS" bateria com a busca por
        # "CUSTO" e devolveria o grupo inteiro em vez do subtotal certo).
        if code_prefix and not code.startswith(code_prefix + '.'):
            continue
        if depth is not None and _depth(code) != depth:
            continue
        nd = normalize_desc(desc)
        if keywords_all and not all(k in nd for k in keywords_all):
            continue
        if keywords_any and not any(k in nd for k in keywords_any):
            continue
        if any(k in nd for k in exclude):
            continue
        val = sat if field == 2 else (sa or 0.0)
        cands.append((code, val))
    if not cands:
        return None, 0.0
    cands.sort(key=lambda t: _depth(t[0]))
    return cands[0]


def resolve_group_codes(bal_rows):
    """Descobre os códigos reais (nessa empresa) dos grandes grupos do
    Balanço e da DRE, procurando pela descrição em vez de assumir um número
    fixo. Sempre tenta um fallback (o número "clássico" 1/2/3/4) se não achar
    nada pela descrição, para nunca quebrar — mas o normal é achar pela
    descrição mesmo, que é o caminho confiável."""
    def only_code(*args, **kw):
        code, _ = find_group_row(*args, **kw)
        return code

    codes = {}
    codes['ativo'] = only_code(bal_rows, keywords_all=['ATIVO'], depth=1) or '1'
    codes['ativo_circ'] = only_code(bal_rows, keywords_all=['ATIVO', 'CIRCULANTE'],
                                     exclude=['NAO CIRCULANTE'], code_prefix=codes['ativo']) or '1.1'
    codes['ativo_nao_circ'] = only_code(bal_rows, keywords_all=['ATIVO', 'NAO CIRCULANTE'],
                                         code_prefix=codes['ativo']) or '1.2'

    codes['passivo'] = only_code(bal_rows, keywords_all=['PASSIVO'], depth=1,
                                  exclude=['CIRCULANTE']) or '2'
    codes['passivo_circ'] = only_code(bal_rows, keywords_all=['PASSIVO', 'CIRCULANTE'],
                                       exclude=['NAO CIRCULANTE'], code_prefix=codes['passivo']) or '2.1'
    codes['passivo_nao_circ'] = only_code(bal_rows, keywords_all=['PASSIVO', 'NAO CIRCULANTE'],
                                           code_prefix=codes['passivo']) or '2.2'
    codes['pl'] = (only_code(bal_rows, keywords_any=['PATRIMONIO LIQUIDO', 'PATRIMONIO NETO'],
                              code_prefix=codes['passivo'])
                   or only_code(bal_rows, keywords_any=['PATRIMONIO LIQUIDO', 'PATRIMONIO NETO'], depth=1)
                   or '2.3')

    codes['despesas'] = only_code(bal_rows, keywords_any=['CUSTO', 'DESPESA'], depth=1) or '3'
    codes['receitas'] = only_code(bal_rows, keywords_any=['RECEITA'], depth=1,
                                   exclude=['DEDUC']) or '4'
    return codes


def _claim_subtotal(bal_rows, group_code, field=2, **kw):
    """Acha o subtotal (por palavra-chave) dentro de um grupo e devolve
    (codigo_achado_ou_None, valor) — o valor vem da mesma linha encontrada,
    não de uma busca separada por código (ver nota em find_group_row sobre
    códigos repetidos). Usado tanto para achar quanto para depois poder
    EXCLUIR esse código de uma busca mais ampla (evita contar 2x)."""
    code, val = find_group_row(bal_rows, code_prefix=group_code, field=field, **kw)
    if not code:
        return None, 0.0
    return code, val


def compute_despesas_operacionais(bal_rows, despesas_group_code, field=2):
    """Quebra as despesas operacionais em categorias por palavra-chave da
    descrição — funciona com qualquer plano de contas. Primeiro tenta achar
    o subtotal já nomeado de cada categoria comum (Comerciais, Despesas com
    Pessoal, Aluguéis, Tributárias/Impostos e Taxas), onde quer que esteja
    dentro do grupo de despesas. Contas-folha que não caem em nenhuma dessas
    categorias nomeadas são classificadas uma a uma pelas mesmas regras de
    palavra-chave (RULES/classify_desc). No final, o que ainda sobrar depois
    de tudo isso cai em "administrativas" como resíduo — calculado contra o
    subtotal REAL de "Despesas Operacionais" do balancete (quando essa linha
    existe), então o TOTAL nunca depende de a categorização ter sido
    perfeita: só a divisão por categoria é que pode não ser exata."""
    # Acha primeiro o subtotal de "Despesas Operacionais" (se existir) para
    # escopar as categorias abaixo SÓ dentro dele — senão uma palavra-chave
    # como "TRIBUTARIA" pode acabar pegando algo do lado de Custos (ex.: ICMS
    # substituição tributária sobre compras, que é custo de mercadoria, não
    # despesa operacional) e desbalancear o resíduo de "administrativas".
    despesas_operacionais_code, despesas_operacionais_val = find_group_row(
        bal_rows, keywords_all=['DESPESA', 'OPERACIONA'], code_prefix=despesas_group_code, field=field)
    despesas_operacionais_real = -despesas_operacionais_val if despesas_operacionais_code else None
    scope_code = despesas_operacionais_code or despesas_group_code

    claimed_codes = []

    def claim(**kw):
        code, val = _claim_subtotal(bal_rows, scope_code, field=field, **kw)
        if code:
            claimed_codes.append(code)
        return -val

    # Despesas financeiras (juros, IOF, tarifas bancárias etc.) costumam estar
    # ANINHADAS dentro de "Despesas Operacionais" no plano de contas (ex.:
    # "3.2.02.05" dentro de "3.2"), mas por convenção contábil NÃO entram no
    # EBITDA — são deduzidas só depois, junto das receitas financeiras. Por
    # isso são "reclamadas" aqui e removidas do subtotal real de despesas
    # operacionais antes de calcular o EBITDA, senão o EBITDA saía subestimado.
    despesas_financeiras = claim(keywords_all=['DESPESA', 'FINANCEIRA'])
    if despesas_operacionais_real is not None:
        despesas_operacionais_real -= despesas_financeiras

    comerciais = claim(keywords_all=['COMERCIAL'])
    pessoal = claim(keywords_all=['PESSOAL'])
    alugueis = claim(keywords_all=['ALUGUEL'])
    tributarias = claim(keywords_all=['TRIBUTARIA'])
    if tributarias == 0.0:
        tributarias = claim(keywords_all=['IMPOSTO', 'TAXA'])

    # contas-folha (sem conta-filha) dentro do escopo que ainda não foram
    # "reclamadas" por nenhum subtotal nomeado acima — classificadas uma a
    # uma pela descrição.
    all_codes = {c for c, _, _, _ in bal_rows}
    cats = {}
    for code, desc, sat, sa in bal_rows:
        if not (code == scope_code or code.startswith(scope_code + '.')):
            continue
        if any(code == cc or code.startswith(cc + '.') for cc in claimed_codes):
            continue
        if any(other.startswith(code + '.') for other in all_codes):
            continue  # não é folha (é um subtotal intermediário) — pula, senão conta 2x
        val = sat if field == 2 else (sa or 0.0)
        cat = classify_desc(desc)
        cats[cat] = cats.get(cat, 0.0) + val

    depreciacao = -cats.get('depreciacao', 0.0)
    transportes = -cats.get('transportes', 0.0)
    manutencao = -cats.get('manutencao', 0.0)
    informatica = -cats.get('informatica', 0.0)
    consultoria = -cats.get('consultoria', 0.0)
    prestacao_servicos = -cats.get('prestacao_servicos', 0.0)
    marketing = -cats.get('marketing', 0.0)
    provisoes_desp = -cats.get('provisoes', 0.0)
    administrativas_leftover = -cats.get('administrativas', 0.0)

    sum_sem_administrativas = (comerciais + pessoal + tributarias + prestacao_servicos + manutencao +
                                alugueis + provisoes_desp + marketing + transportes + consultoria +
                                informatica + depreciacao)

    if despesas_operacionais_real is not None:
        administrativas = despesas_operacionais_real - sum_sem_administrativas
        despesas_operacionais_sem_depre = despesas_operacionais_real - depreciacao
    else:
        administrativas = administrativas_leftover
        despesas_operacionais_sem_depre = sum_sem_administrativas + administrativas_leftover - depreciacao

    return {
        'comerciais': comerciais, 'administrativas': administrativas, 'pessoal': pessoal,
        'tributarias': tributarias, 'prestacao_servicos': prestacao_servicos, 'manutencao': manutencao,
        'alugueis': alugueis, 'provisoes_desp': provisoes_desp, 'marketing': marketing,
        'transportes': transportes, 'consultoria': consultoria, 'informatica': informatica,
        'depreciacao': depreciacao, 'despesas_operacionais_sem_depre': despesas_operacionais_sem_depre,
        'despesas_financeiras': despesas_financeiras,
    }


# --------------------------------------------------------------------------
# 3. Cálculo do Balanço detalhado (por empresa, por período)
# --------------------------------------------------------------------------

def compute_bp_detalhado(bal_rows, field=2, codes=None):
    codes = codes or resolve_group_codes(bal_rows)
    c_ativo = codes['ativo']; c_ativo_circ = codes['ativo_circ']; c_ativo_nao_circ = codes['ativo_nao_circ']
    c_passivo = codes['passivo']; c_passivo_circ = codes['passivo_circ']; c_passivo_nao_circ = codes['passivo_nao_circ']
    c_pl = codes['pl']; c_despesas = codes['despesas']; c_receitas = codes['receitas']

    g = lambda code: get_code_value(bal_rows, code, field=field)
    fl = lambda prefix, kws, **kw: find_leaf_by_desc_field(bal_rows, prefix, kws, field=field, **kw)

    # Cada linha de detalhe é achada por palavra-chave da descrição (não por
    # código fixo), dentro do grupo já identificado por resolve_group_codes.
    # O que não bater com nenhuma palavra-chave cai no "residual" (outros_*),
    # calculado contra o TOTAL REAL do grupo (linha de subtotal do próprio
    # balancete) — assim o total do Balanço nunca depende de a categorização
    # de detalhe ter sido perfeita, só a divisão em sub-linhas é que pode não
    # ser exata para planos de conta muito diferentes do padrão.
    caixa = fl(c_ativo_circ, ['CAIXA']) + fl(c_ativo_circ, ['BANCO']) + fl(c_ativo_circ, ['APLICACAO'])
    clientes = fl(c_ativo_circ, ['CLIENTE']) + fl(c_ativo_circ, ['DUPLICATA'])
    estoques = fl(c_ativo_circ, ['ESTOQUE'])
    adiantamentos = fl(c_ativo_circ, ['ADIANTAMENTO'])
    ativo_circ_real = g(c_ativo_circ)
    outros_creditos_circ = ativo_circ_real - (caixa + clientes + estoques + adiantamentos)

    realizavel_lp = fl(c_ativo_nao_circ, ['REALIZAVEL'])
    investimentos = fl(c_ativo_nao_circ, ['INVESTIMENTO'])
    imobilizado = fl(c_ativo_nao_circ, ['IMOBILIZADO'])
    intangivel = fl(c_ativo_nao_circ, ['INTANGIVEL'])
    ativo_nao_circ_real = g(c_ativo_nao_circ)
    realizavel_lp += ativo_nao_circ_real - (realizavel_lp + investimentos + imobilizado + intangivel)

    ativo_total = g(c_ativo)

    fornecedores = fl(c_passivo_circ, ['FORNECEDOR'])
    impostos_recolher = fl(c_passivo_circ, ['IMPOSTO'])
    obr_pessoal = fl(c_passivo_circ, ['PESSOAL']) + fl(c_passivo_circ, ['TRABALHISTA'])
    emprestimos_cp = fl(c_passivo_circ, ['EMPRESTIMO']) + fl(c_passivo_circ, ['FINANCIAMENTO'])
    provisoes_circ = fl(c_passivo_circ, ['PROVIS'])
    passivo_circ_real = g(c_passivo_circ)
    outras_obrig_circ = passivo_circ_real - fornecedores - impostos_recolher - obr_pessoal - emprestimos_cp - provisoes_circ

    emprestimos_lp = fl(c_passivo_nao_circ, ['FINANCIAMENTO']) + fl(c_passivo_nao_circ, ['EMPRESTIMO']) + \
                      fl(c_passivo_nao_circ, ['ARRENDAMENTO']) + fl(c_passivo_nao_circ, ['LEASING'])
    passivo_nao_circ_real = g(c_passivo_nao_circ)
    outras_obrig_nao_circ = passivo_nao_circ_real - emprestimos_lp

    capital_social = fl(c_pl, ['CAPITAL SOCIAL'])
    reservas = fl(c_pl, ['RESERVA'])
    lucros_prejuizos = fl(c_pl, ['LUCRO']) + fl(c_pl, ['PREJUIZO'])
    if field == 2:
        resultado_periodo = -(get_code_value(bal_rows, c_despesas, field=2) + get_code_value(bal_rows, c_receitas, field=2))
    else:
        resultado_periodo = 0.0
    pl_real = g(c_pl)
    outros_pl = pl_real - (capital_social + reservas + lucros_prejuizos)
    pl_total = pl_real - resultado_periodo

    passivo_pl_total = passivo_circ_real + passivo_nao_circ_real + pl_total

    return {
        'caixa': caixa, 'clientes': clientes, 'estoques': estoques,
        'adiantamentos': adiantamentos, 'outros_creditos_circ': outros_creditos_circ,
        'ativo_circ': ativo_circ_real,
        'realizavel_lp': realizavel_lp, 'investimentos': investimentos,
        'imobilizado': imobilizado, 'intangivel': intangivel,
        'ativo_nao_circ': ativo_nao_circ_real, 'ativo_total': ativo_total,
        'fornecedores': fornecedores, 'impostos_recolher': impostos_recolher,
        'obr_pessoal': obr_pessoal, 'emprestimos_cp': emprestimos_cp,
        'outras_obrig_circ': outras_obrig_circ, 'provisoes_circ': provisoes_circ,
        'passivo_circ': passivo_circ_real,
        'emprestimos_lp': emprestimos_lp, 'outras_obrig_nao_circ': outras_obrig_nao_circ,
        'passivo_nao_circ': passivo_nao_circ_real,
        'capital_social': capital_social, 'reservas': reservas,
        'lucros_prejuizos': lucros_prejuizos, 'resultado_periodo': resultado_periodo,
        'outros_pl': outros_pl, 'pl_total': pl_total,
        'passivo_pl_total': passivo_pl_total,
    }


BP_ROWS = [
    ("ATIVO", 0, 'ativo_total'),
    ("Ativo Circulante", 1, 'ativo_circ'),
    ("Caixa e equivalentes de caixa", 2, 'caixa'),
    ("Clientes", 2, 'clientes'),
    ("Estoques", 2, 'estoques'),
    ("Adiantamentos", 2, 'adiantamentos'),
    ("Outros Créditos (circulante)", 2, 'outros_creditos_circ'),
    ("Ativo Não Circulante", 1, 'ativo_nao_circ'),
    ("Realizável a longo prazo", 2, 'realizavel_lp'),
    ("Investimentos", 2, 'investimentos'),
    ("Imobilizado (líquido)", 2, 'imobilizado'),
    ("Intangível", 2, 'intangivel'),
    ("TOTAL DO ATIVO", 0, 'ativo_total'),
    ("PASSIVO CIRCULANTE + NÃO CIRCULANTE", 0, 'passivo_pl_total'),
    ("Passivo Circulante", 1, 'passivo_circ'),
    ("Fornecedores", 2, 'fornecedores'),
    ("Impostos, taxas e contribuições a recolher", 2, 'impostos_recolher'),
    ("Obrigações sociais e trabalhistas", 2, 'obr_pessoal'),
    ("Empréstimos e financiamentos (curto prazo)", 2, 'emprestimos_cp'),
    ("Outras obrigações", 2, 'outras_obrig_circ'),
    ("Provisões", 2, 'provisoes_circ'),
    ("Passivo Não Circulante", 1, 'passivo_nao_circ'),
    ("Empréstimos e financiamentos (longo prazo)", 2, 'emprestimos_lp'),
    ("Outras obrigações (não circulante)", 2, 'outras_obrig_nao_circ'),
    ("Patrimônio Líquido", 1, 'pl_total'),
    ("Capital social", 2, 'capital_social'),
    ("Reservas de lucro / subvenção", 2, 'reservas'),
    ("Lucros ou prejuízos acumulados (exercícios anteriores)", 2, 'lucros_prejuizos'),
    ("Resultado do exercício em curso (acumulado no período)", 2, 'resultado_periodo'),
    ("Outras contas do Patrimônio Líquido", 2, 'outros_pl'),
    ("TOTAL DO PASSIVO + PATRIMÔNIO LÍQUIDO", 0, 'passivo_pl_total'),
]


# --------------------------------------------------------------------------
# 4. Cálculo da DRE detalhada
# --------------------------------------------------------------------------

def compute_dre_detalhado(bal_rows, codes=None):
    codes = codes or resolve_group_codes(bal_rows)
    c_despesas = codes['despesas']; c_receitas = codes['receitas']
    g = lambda code: get_code_value(bal_rows, code)

    _, receita_bruta_raw = _claim_subtotal(bal_rows, c_receitas, keywords_all=['RECEITA', 'BRUTA'], exclude=['DEDUC'])
    _, deducoes_raw = _claim_subtotal(bal_rows, c_receitas,
                                       keywords_any=['DEDUCAO DA RECEITA', 'DEDUCOES DA RECEITA',
                                                      'IMPOSTO SOBRE VENDA', 'IMPOSTOS SOBRE VENDA'])
    receita_bruta = -receita_bruta_raw
    deducoes = -deducoes_raw
    receita_liquida = receita_bruta + deducoes

    _, custos_raw = _claim_subtotal(bal_rows, c_despesas, keywords_all=['CUSTO'])
    custos = -custos_raw
    lucro_bruto = receita_liquida + custos

    desp = compute_despesas_operacionais(bal_rows, c_despesas)
    comerciais = desp['comerciais']; administrativas = desp['administrativas']; pessoal = desp['pessoal']
    tributarias = desp['tributarias']; prestacao_servicos = desp['prestacao_servicos']
    manutencao = desp['manutencao']; alugueis = desp['alugueis']; provisoes_desp = desp['provisoes_desp']
    marketing = desp['marketing']; transportes = desp['transportes']; consultoria = desp['consultoria']
    informatica = desp['informatica']; depreciacao = desp['depreciacao']
    despesas_operacionais_sem_depre = desp['despesas_operacionais_sem_depre']
    despesas_financeiras = desp['despesas_financeiras']

    ebitda = lucro_bruto + despesas_operacionais_sem_depre
    ebit = ebitda + depreciacao

    _, receitas_financeiras_raw = _claim_subtotal(bal_rows, c_receitas, keywords_all=['RECEITA', 'FINANCEIRA'])
    receitas_financeiras = -receitas_financeiras_raw

    resultado_liquido = -(g(c_despesas) + g(c_receitas))
    outras_nao_operacionais = resultado_liquido - (ebit + despesas_financeiras + receitas_financeiras)

    return {
        'receita_bruta': receita_bruta, 'deducoes': deducoes, 'receita_liquida': receita_liquida,
        'custos': custos, 'lucro_bruto': lucro_bruto,
        'comerciais': comerciais, 'administrativas': administrativas, 'pessoal': pessoal,
        'tributarias': tributarias, 'prestacao_servicos': prestacao_servicos, 'manutencao': manutencao,
        'alugueis': alugueis, 'provisoes_desp': provisoes_desp, 'marketing': marketing,
        'transportes': transportes, 'consultoria': consultoria, 'informatica': informatica,
        'despesas_operacionais_sem_depre': despesas_operacionais_sem_depre,
        'ebitda': ebitda, 'depreciacao': depreciacao, 'ebit': ebit,
        'despesas_financeiras': despesas_financeiras, 'receitas_financeiras': receitas_financeiras,
        'outras_nao_operacionais': outras_nao_operacionais,
        'resultado_liquido': resultado_liquido,
    }


DRE_ROWS = [
    ("Receita Bruta", 0, 'receita_bruta'),
    ("(-) Deduções da receita bruta", 1, 'deducoes'),
    ("Receita Líquida", 0, 'receita_liquida'),
    ("(-) Custos das mercadorias/serviços", 1, 'custos'),
    ("Lucro Bruto", 0, 'lucro_bruto'),
    ("Despesas Operacionais", 0, 'despesas_operacionais_sem_depre'),
    ("Comerciais", 1, 'comerciais'),
    ("Administrativas", 1, 'administrativas'),
    ("Com pessoal", 1, 'pessoal'),
    ("Tributárias", 1, 'tributarias'),
    ("Prestação de serviços", 1, 'prestacao_servicos'),
    ("Manutenção", 1, 'manutencao'),
    ("Aluguéis", 1, 'alugueis'),
    ("Provisões", 1, 'provisoes_desp'),
    ("Marketing", 1, 'marketing'),
    ("Transportes", 1, 'transportes'),
    ("Consultoria técnica", 1, 'consultoria'),
    ("Informática", 1, 'informatica'),
    ("EBITDA", 0, 'ebitda'),
    ("(-) Depreciação e amortização", 1, 'depreciacao'),
    ("EBIT", 0, 'ebit'),
    ("(-) Despesas financeiras", 1, 'despesas_financeiras'),
    ("(+) Receitas financeiras", 1, 'receitas_financeiras'),
    ("Outras receitas/despesas não operacionais", 1, 'outras_nao_operacionais'),
    ("RESULTADO LÍQUIDO DO PERÍODO", 0, 'resultado_liquido'),
]


# --------------------------------------------------------------------------
# 4b. DFC detalhada (método indireto)
# --------------------------------------------------------------------------

PARTES_RELACIONADAS_KEYWORDS = ['LIGADA', 'LIGADO', 'RELACIONAD', 'SOCIO']


def find_partes_relacionadas(bal_rows, code_prefix, field):
    total = 0.0
    for kw in PARTES_RELACIONADAS_KEYWORDS:
        val = find_leaf_by_desc_field(bal_rows, code_prefix, [kw], field=field)
        if val != 0.0:
            total += val
    return total


def compute_dfc_detalhado(bal_rows):
    codes = resolve_group_codes(bal_rows)
    atual = compute_bp_detalhado(bal_rows, field=2, codes=codes)
    ant = compute_bp_detalhado(bal_rows, field=3, codes=codes)
    dre = compute_dre_detalhado(bal_rows, codes=codes)

    def var(key):
        return ant[key] - atual[key]

    lucro = atual['resultado_periodo']
    depreciacao_addback = -dre['depreciacao']

    titulos_receber_lp_atual = find_leaf_by_desc_field(bal_rows, codes['ativo_nao_circ'], ['TITULOS A RECEBER'], field=2)
    titulos_receber_lp_ant = find_leaf_by_desc_field(bal_rows, codes['ativo_nao_circ'], ['TITULOS A RECEBER'], field=3)
    op_titulos_receber_lp = titulos_receber_lp_ant - titulos_receber_lp_atual

    pr_ativo_atual = find_partes_relacionadas(bal_rows, codes['ativo_circ'], field=2)
    pr_ativo_ant = find_partes_relacionadas(bal_rows, codes['ativo_circ'], field=3)
    pr_pas_21_atual = find_partes_relacionadas(bal_rows, codes['passivo_circ'], field=2)
    pr_pas_21_ant = find_partes_relacionadas(bal_rows, codes['passivo_circ'], field=3)
    pr_pas_22_atual = find_partes_relacionadas(bal_rows, codes['passivo_nao_circ'], field=2)
    pr_pas_22_ant = find_partes_relacionadas(bal_rows, codes['passivo_nao_circ'], field=3)
    var_partes_relacionadas = ((pr_ativo_ant - pr_ativo_atual) + (pr_pas_21_ant - pr_pas_21_atual) +
                               (pr_pas_22_ant - pr_pas_22_atual))

    op_clientes = var('clientes')
    op_estoques = var('estoques')
    op_adiantamentos = var('adiantamentos')
    op_outros_creditos = var('outros_creditos_circ') - (pr_ativo_ant - pr_ativo_atual)
    op_fornecedores = var('fornecedores')
    op_impostos = var('impostos_recolher')
    op_obr_pessoal = var('obr_pessoal')
    op_provisoes = var('provisoes_circ')
    op_outras_obrig = var('outras_obrig_circ')
    subtotal_operacional = (lucro + depreciacao_addback + op_clientes + op_estoques + op_adiantamentos +
                             op_outros_creditos + op_titulos_receber_lp + op_fornecedores + op_impostos +
                             op_obr_pessoal + op_provisoes + op_outras_obrig + var_partes_relacionadas)

    var_imob_intang_bruto = var('imobilizado') + var('intangivel')
    imob_intang_liquido = var_imob_intang_bruto - depreciacao_addback

    var_realizavel_lp = var('realizavel_lp') - op_titulos_receber_lp
    var_investimentos = var('investimentos')
    subtotal_investimento = var_realizavel_lp + var_investimentos + imob_intang_liquido

    var_emprestimos_cp = var('emprestimos_cp') - (pr_pas_21_ant - pr_pas_21_atual)
    var_emprestimos_lp = var('emprestimos_lp')
    var_outras_obrig_nc = var('outras_obrig_nao_circ') - (pr_pas_22_ant - pr_pas_22_atual)
    var_pl_ex_resultado = var('capital_social') + var('reservas') + var('lucros_prejuizos') + var('outros_pl')
    subtotal_financiamento = var_emprestimos_cp + var_emprestimos_lp + var_outras_obrig_nc + var_pl_ex_resultado

    variacao_caixa = subtotal_operacional + subtotal_investimento + subtotal_financiamento
    caixa_inicial = ant['caixa']
    caixa_final = atual['caixa']

    return {
        'lucro': lucro, 'depreciacao_addback': depreciacao_addback,
        'op_clientes': op_clientes, 'op_estoques': op_estoques, 'op_adiantamentos': op_adiantamentos,
        'op_outros_creditos': op_outros_creditos, 'op_titulos_receber_lp': op_titulos_receber_lp,
        'op_fornecedores': op_fornecedores,
        'op_impostos': op_impostos, 'op_obr_pessoal': op_obr_pessoal, 'op_provisoes': op_provisoes,
        'op_outras_obrig': op_outras_obrig, 'subtotal_operacional': subtotal_operacional,
        'var_realizavel_lp': var_realizavel_lp, 'var_investimentos': var_investimentos,
        'imob_intang_liquido': imob_intang_liquido, 'subtotal_investimento': subtotal_investimento,
        'var_emprestimos_cp': var_emprestimos_cp, 'var_emprestimos_lp': var_emprestimos_lp,
        'var_outras_obrig_nc': var_outras_obrig_nc, 'var_pl_ex_resultado': var_pl_ex_resultado,
        'var_partes_relacionadas': var_partes_relacionadas,
        'subtotal_financiamento': subtotal_financiamento,
        'variacao_caixa': variacao_caixa, 'caixa_inicial': caixa_inicial, 'caixa_final': caixa_final,
    }


DFC_ROWS = [
    ("FLUXO DE CAIXA DAS ATIVIDADES OPERACIONAIS", 0, None),
    ("Lucro líquido do período", 1, 'lucro'),
    ("(+) Depreciação e amortização", 1, 'depreciacao_addback'),
    ("Clientes", 1, 'op_clientes'),
    ("Estoques", 1, 'op_estoques'),
    ("Adiantamentos", 1, 'op_adiantamentos'),
    ("Outros Créditos (circulante)", 1, 'op_outros_creditos'),
    ("Fornecedores", 1, 'op_fornecedores'),
    ("Impostos, taxas e contribuições a recolher", 1, 'op_impostos'),
    ("Obrigações sociais e trabalhistas", 1, 'op_obr_pessoal'),
    ("Provisões", 1, 'op_provisoes'),
    ("Outras contas a receber (Realizável a Longo Prazo)", 1, 'op_titulos_receber_lp'),
    ("Outras obrigações (circulante)", 1, 'op_outras_obrig'),
    ("Empréstimos/créditos com sócios e partes relacionadas", 1, 'var_partes_relacionadas'),
    ("CAIXA GERADO (CONSUMIDO) PELAS ATIVIDADES OPERACIONAIS", 0, 'subtotal_operacional'),
    ("FLUXO DE CAIXA DAS ATIVIDADES DE INVESTIMENTO", 0, None),
    ("Realizável a longo prazo", 1, 'var_realizavel_lp'),
    ("Investimentos", 1, 'var_investimentos'),
    ("Imobilizado e Intangível (líquido de depreciação/amortização)", 1, 'imob_intang_liquido'),
    ("CAIXA GERADO (CONSUMIDO) PELAS ATIVIDADES DE INVESTIMENTO", 0, 'subtotal_investimento'),
    ("FLUXO DE CAIXA DAS ATIVIDADES DE FINANCIAMENTO", 0, None),
    ("Empréstimos e financiamentos (curto prazo)", 1, 'var_emprestimos_cp'),
    ("Empréstimos e financiamentos (longo prazo)", 1, 'var_emprestimos_lp'),
    ("Outras obrigações (não circulante)", 1, 'var_outras_obrig_nc'),
    ("Patrimônio Líquido (aportes/distribuições/reservas)", 1, 'var_pl_ex_resultado'),
    ("CAIXA GERADO (CONSUMIDO) PELAS ATIVIDADES DE FINANCIAMENTO", 0, 'subtotal_financiamento'),
    ("VARIAÇÃO DE CAIXA E EQUIVALENTES NO PERÍODO", 0, 'variacao_caixa'),
    ("(+) Caixa e equivalentes no início do período", 1, 'caixa_inicial'),
    ("(=) Caixa e equivalentes no final do período", 0, 'caixa_final'),
]


def sum_dicts(dict_list):
    out = {}
    for d in dict_list:
        for k, v in d.items():
            out[k] = out.get(k, 0.0) + v
    return out


# --------------------------------------------------------------------------
# 5. Orquestração: de balancetes brutos a bp/dre/dfc consolidados por período
# --------------------------------------------------------------------------

def consolidate(companies, n_periodos=None):
    """companies: dict cnpj -> {'nome':..., 'periodos': [(periodo, bal_rows, fname), ...]}
    Retorna (bp_consolidado, dre_consolidado, dfc_consolidado, period_labels_used, warnings)
    onde cada *_consolidado é uma lista com um dict por período + 1 dict "agregado" no final."""
    if n_periodos is None:
        n_periodos = max((len(d['periodos']) for d in companies.values()), default=0)
    warnings = []
    incompletas = [d['nome'] for d in companies.values() if len(d['periodos']) != n_periodos]
    if incompletas:
        warnings.append(f"Empresas com número de períodos diferente do esperado ({n_periodos}): {', '.join(map(str, incompletas))}")

    bp_by_period = [[] for _ in range(n_periodos)]
    dre_by_period = [[] for _ in range(n_periodos)]
    dfc_by_period = [[] for _ in range(n_periodos)]
    period_labels_seen = [None] * n_periodos
    for cnpj, info in companies.items():
        periodos_ordenados = sorted(info['periodos'], key=lambda p: (p[0] or ""))
        for idx, (periodo, rows, fname) in enumerate(periodos_ordenados[:n_periodos]):
            bp_by_period[idx].append(compute_bp_detalhado(rows))
            dre_by_period[idx].append(compute_dre_detalhado(rows))
            dfc_by_period[idx].append(compute_dfc_detalhado(rows))
            if period_labels_seen[idx] is None and periodo:
                period_labels_seen[idx] = periodo

    bp_periods = [sum_dicts(p) for p in bp_by_period]
    dre_periods = [sum_dicts(p) for p in dre_by_period]
    dfc_periods = [sum_dicts(p) for p in dfc_by_period]

    bp_semestre = dict(bp_periods[-1]) if bp_periods else {}
    dre_semestre = sum_dicts(dre_periods)
    dfc_semestre = sum_dicts(dfc_periods)
    if dfc_periods:
        dfc_semestre['caixa_inicial'] = dfc_periods[0]['caixa_inicial']
        dfc_semestre['caixa_final'] = dfc_periods[-1]['caixa_final']

    bp_consolidado = bp_periods + [bp_semestre]
    dre_consolidado = dre_periods + [dre_semestre]
    dfc_consolidado = dfc_periods + [dfc_semestre]

    # checagem de consistência
    checks = []
    resultado_bp_ref = [p['resultado_periodo'] for p in bp_periods]
    resultado_bp_ref.append(sum(resultado_bp_ref))
    labels_chk = [f"Período {i+1}" for i in range(n_periodos)] + ["Agregado"]
    for idx in range(len(bp_consolidado)):
        diff_bp = bp_consolidado[idx]['ativo_total'] + bp_consolidado[idx]['passivo_pl_total']
        diff_dre = dre_consolidado[idx]['resultado_liquido'] - resultado_bp_ref[idx]
        diff_dfc = dfc_consolidado[idx]['variacao_caixa'] - (dfc_consolidado[idx]['caixa_final'] - dfc_consolidado[idx]['caixa_inicial'])
        checks.append((labels_chk[idx], diff_bp, diff_dre, diff_dfc))
        if max(abs(diff_bp), abs(diff_dre), abs(diff_dfc)) > 1.0:
            warnings.append(f"{labels_chk[idx]}: diferença de consistência acima de R$1 (BP={diff_bp:.2f}, DRE={diff_dre:.2f}, DFC={diff_dfc:.2f}) — confira códigos contábeis não mapeados.")

    period_labels = [lbl or f"Período {i+1}" for i, lbl in enumerate(period_labels_seen)] + ["Agregado (soma dos períodos)"]

    return {
        'bp': bp_consolidado, 'dre': dre_consolidado, 'dfc': dfc_consolidado,
        'bp_rows': BP_ROWS, 'dre_rows': DRE_ROWS, 'dfc_rows': DFC_ROWS,
        'period_labels': period_labels,
        'checks': checks, 'warnings': warnings,
        'n_periodos': n_periodos, 'n_empresas': len(companies),
    }


# --------------------------------------------------------------------------
# 6. Indicadores financeiros a partir do resultado de consolidate()
# --------------------------------------------------------------------------

def mi(v):
    return v / 1_000_000


def compute_indicators(statements):
    bp, dre, dfc = statements['bp'], statements['dre'], statements['dfc']
    n = len(bp)
    out = []
    for i in range(n):
        ativo_total = bp[i]['ativo_total']
        ativo_circ = bp[i]['ativo_circ']
        passivo_circ = -bp[i]['passivo_circ']
        passivo_nao_circ = -bp[i]['passivo_nao_circ']
        pl_total = -bp[i]['pl_total']
        emprestimos_cp = -bp[i]['emprestimos_cp']
        emprestimos_lp = -bp[i]['emprestimos_lp']
        caixa = bp[i]['caixa']
        divida_bruta = emprestimos_cp + emprestimos_lp
        divida_liquida = divida_bruta - caixa

        receita_liquida = dre[i]['receita_liquida']
        lucro_bruto = dre[i]['lucro_bruto']
        ebitda = dre[i]['ebitda']
        resultado_liquido = dre[i]['resultado_liquido']
        despesas_financeiras = -dre[i]['despesas_financeiras']

        def safe_div(a, b):
            return a / b if b else float('nan')

        out.append({
            'ativo_total_mi': mi(ativo_total), 'ativo_circ_mi': mi(ativo_circ),
            'passivo_circ_mi': mi(passivo_circ), 'passivo_nao_circ_mi': mi(passivo_nao_circ),
            'pl_total_mi': mi(pl_total),
            'emprestimos_cp_mi': mi(emprestimos_cp), 'emprestimos_lp_mi': mi(emprestimos_lp),
            'caixa_mi': mi(caixa), 'divida_bruta_mi': mi(divida_bruta), 'divida_liquida_mi': mi(divida_liquida),
            'receita_liquida_mi': mi(receita_liquida), 'lucro_bruto_mi': mi(lucro_bruto),
            'ebitda_mi': mi(ebitda), 'resultado_liquido_mi': mi(resultado_liquido),
            'despesas_financeiras_mi': mi(despesas_financeiras),
            'liquidez_corrente': safe_div(ativo_circ, passivo_circ),
            'margem_bruta_pct': safe_div(lucro_bruto, receita_liquida) * 100,
            'margem_ebitda_pct': safe_div(ebitda, receita_liquida) * 100,
            'margem_liquida_pct': safe_div(resultado_liquido, receita_liquida) * 100,
            'endividamento_geral_pct': safe_div(passivo_circ + passivo_nao_circ, ativo_total) * 100,
            'pl_sobre_ativo_pct': safe_div(pl_total, ativo_total) * 100,
            'concentracao_divida_cp_pct': safe_div(passivo_circ, passivo_circ + passivo_nao_circ) * 100,
            'cobertura_juros_x': safe_div(ebitda, despesas_financeiras),
            'caixa_operacional_mi': mi(dfc[i]['subtotal_operacional']),
            'caixa_investimento_mi': mi(dfc[i]['subtotal_investimento']),
            'caixa_financiamento_mi': mi(dfc[i]['subtotal_financiamento']),
            'variacao_caixa_mi': mi(dfc[i]['variacao_caixa']),
            'caixa_inicial_mi': mi(dfc[i]['caixa_inicial']), 'caixa_final_mi': mi(dfc[i]['caixa_final']),
            'op_clientes_mi': mi(dfc[i]['op_clientes']), 'op_estoques_mi': mi(dfc[i]['op_estoques']),
            'op_fornecedores_mi': mi(dfc[i]['op_fornecedores']),
        })

    # alavancagem/ROE/ROA anualizados: usa o último período de fluxo (penúltimo
    # índice, já que o último é o agregado) e o agregado para o fator de
    # anualização (ex.: semestre -> x2 para virar "anual").
    last_flow_idx = n - 2 if n >= 2 else n - 1
    agg_idx = n - 1
    fator_anual = statements['n_periodos'] * (12 / max(statements['n_periodos'], 1)) / statements['n_periodos'] * statements['n_periodos']
    # fator simples: se n_periodos trimestres cobertos, anualizar = (4 / n_periodos)
    fator_anual = 4.0 / max(statements['n_periodos'], 1)
    ebitda_anualizado = out[agg_idx]['ebitda_mi'] * fator_anual
    resultado_anualizado = out[agg_idx]['resultado_liquido_mi'] * fator_anual

    indicators = {
        'por_periodo': out,
        'alavancagem_x': out[last_flow_idx]['divida_liquida_mi'] / ebitda_anualizado if ebitda_anualizado else float('nan'),
        'roe_anualizado_pct': resultado_anualizado / out[last_flow_idx]['pl_total_mi'] * 100 if out[last_flow_idx]['pl_total_mi'] else float('nan'),
        'roa_anualizado_pct': resultado_anualizado / out[last_flow_idx]['ativo_total_mi'] * 100 if out[last_flow_idx]['ativo_total_mi'] else float('nan'),
        'fator_anualizacao': fator_anual,
        'last_flow_idx': last_flow_idx, 'agg_idx': agg_idx,
    }
    if n >= 2:
        p0, p1 = out[0], out[n - 2] if n > 2 else out[0]
        # crescimento entre o primeiro e o último período de fluxo (não o agregado)
        first, last = out[0], out[max(n - 2, 0)]
        def growth(key):
            return (last[key] - first[key]) / first[key] * 100 if first[key] else float('nan')
        indicators['crescimento_pct'] = {
            'receita_liquida': growth('receita_liquida_mi'),
            'ebitda': growth('ebitda_mi'),
            'resultado_liquido': growth('resultado_liquido_mi'),
        }
    return indicators


# --------------------------------------------------------------------------
# 6b. Variação período a período das contas da DRE (para comentário do usuário)
# --------------------------------------------------------------------------

def build_dre_variacao_pares(statements, indicators, labels, pct_threshold=20.0, valor_threshold_mi=0.5):
    """Compara cada período de fluxo com o período seguinte (período a período,
    não só primeiro x último) e marca como 'relevante' toda conta da DRE cuja
    variação ultrapasse o limite percentual OU o limite em R$ milhões — o que
    disparar primeiro. Serve tanto para a UI do app (pedir comentário do
    usuário) quanto para o PPT (mesmos números, mesma marcação)."""
    dre = statements["dre"]
    dre_rows = statements["dre_rows"]
    last_flow_idx = indicators["last_flow_idx"]

    pares = []
    for i in range(last_flow_idx):
        j = i + 1
        linhas = []
        for label, depth, key in dre_rows:
            if key is None:
                continue
            v0 = dre[i].get(key, 0.0) / 1_000_000.0
            v1 = dre[j].get(key, 0.0) / 1_000_000.0
            diff = v1 - v0
            pct = (diff / abs(v0) * 100.0) if v0 else float("nan")
            pct_ok = (pct == pct) and abs(pct) >= pct_threshold  # pct==pct descarta NaN
            relevante = pct_ok or abs(diff) >= valor_threshold_mi
            linhas.append({
                "label": label, "depth": depth, "key": key,
                "v0": v0, "v1": v1, "diff": diff, "pct": pct, "relevante": relevante,
            })
        pares.append({
            "idx_a": i, "idx_b": j,
            "label_a": labels[i], "label_b": labels[j],
            "linhas": linhas,
        })
    return pares


# --------------------------------------------------------------------------
# 7. Escrita do Excel consolidado (mesmo layout usado nas entregas anteriores)
# --------------------------------------------------------------------------

NOTE_BP = ("Linhas em itálico foram adicionadas ao modelo original para não esconder valores "
           "materiais que não se encaixam nas contas específicas do balancete de cada empresa. "
           "A coluna \"Agregado\" do Balanço é a MESMA posição do último período incluído "
           "(Balanço é uma fotografia num instante, não soma de períodos).")
NOTE_DRE = ("\"Despesas Gerais\" do balancete foi separada por palavra-chave da descrição de cada "
            "conta. \"Administrativas\" é a categoria residual. A coluna \"Agregado\" soma todos os "
            "períodos incluídos (DRE é demonstração de fluxo).")
NOTE_DFC = ("Método indireto, a partir do Saldo Anterior x Saldo Atual de cada conta do balancete. "
            "A coluna \"Agregado\" soma todos os períodos incluídos; Caixa Inicial/Final não são "
            "somados (são saldos, não fluxos).")


def write_two_col_sheet(ws, title, subtitle, note, row_template, values_by_period, labels):
    ncols = len(values_by_period)
    last_col = 1 + ncols
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, title).font = Font(name=FONT_NAME, size=14, bold=True, color="1F4E78")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1, subtitle).font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    note_cell = ws.cell(3, 1, note)
    note_cell.font = Font(name=FONT_NAME, size=8, italic=True, color="833C0C")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 50

    header_row = 5
    ws.cell(header_row, 1, "Descrição").font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    for j, lbl in enumerate(labels):
        cell = ws.cell(header_row, 2 + j, lbl)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    for col in range(1, last_col + 1):
        ws.cell(header_row, col).fill = HEADER_FILL
        ws.cell(header_row, col).border = BORDER
    ws.row_dimensions[header_row].height = 30

    r = header_row + 1
    for label, depth, key in row_template:
        desc_cell = ws.cell(r, 1, label)
        if key is None:
            desc_cell.font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E78")
            for col in range(1, last_col + 1):
                ws.cell(r, col).fill = SUBHEADER_FILL
                ws.cell(r, col).border = BORDER
            r += 1
            continue
        desc_cell.font = Font(name=FONT_NAME, bold=(depth == 0), size=10 if depth == 0 else 9)
        desc_cell.alignment = Alignment(indent=depth * 2)
        for j in range(ncols):
            val = values_by_period[j].get(key, 0.0)
            cell = ws.cell(r, 2 + j, val)
            cell.number_format = NUMFMT
            cell.font = Font(name=FONT_NAME, size=9, bold=(depth == 0))
        if depth == 0:
            for col in range(1, last_col + 1):
                ws.cell(r, col).fill = TOTAL_FILL
        elif depth == 1:
            for col in range(1, last_col + 1):
                ws.cell(r, col).fill = SUBHEADER_FILL
        for col in range(1, last_col + 1):
            ws.cell(r, col).border = BORDER
        r += 1

    ws.freeze_panes = ws.cell(header_row + 1, 2)
    ws.column_dimensions['A'].width = 52
    for j in range(ncols):
        ws.column_dimensions[get_column_letter(2 + j)].width = 20


def build_excel(statements, out_path, recalc=True, group_name="Empresa/Grupo"):
    subtitle = ("Consolidação por soma simples das empresas informadas (sem eliminação de saldos "
                "intercompanhias, se forem mais de uma empresa do mesmo grupo)." if statements.get('n_empresas', 1) != 1
                else "Demonstrativo consolidado por período.")
    wb = Workbook()
    ws_bp = wb.active
    ws_bp.title = "Balanço Consolidado"
    write_two_col_sheet(ws_bp, f"BALANÇO PATRIMONIAL CONSOLIDADO — {group_name.upper()}",
                         subtitle,
                         NOTE_BP, statements['bp_rows'], statements['bp'], statements['period_labels'])
    ws_dre = wb.create_sheet("DRE Consolidada")
    write_two_col_sheet(ws_dre, f"DEMONSTRAÇÃO DO RESULTADO CONSOLIDADA — {group_name.upper()}",
                         subtitle,
                         NOTE_DRE, statements['dre_rows'], statements['dre'], statements['period_labels'])
    ws_dfc = wb.create_sheet("DFC Consolidada")
    write_two_col_sheet(ws_dfc, f"DEMONSTRAÇÃO DOS FLUXOS DE CAIXA CONSOLIDADA — {group_name.upper()}",
                         subtitle + " Método indireto.",
                         NOTE_DFC, statements['dfc_rows'], statements['dfc'], statements['period_labels'])
    wb.save(out_path)

    if recalc:
        try:
            outdir = tempfile.mkdtemp(prefix="recalc_det_")
            subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", outdir, out_path],
                           capture_output=True, check=True, timeout=120)
            recalculated = os.path.join(outdir, os.path.basename(out_path))
            if os.path.exists(recalculated):
                shutil.move(recalculated, out_path)
        except Exception:
            pass  # se o soffice não estiver disponível, entrega sem recálculo de fórmulas (não há fórmulas aqui, só valores)
    return out_path
